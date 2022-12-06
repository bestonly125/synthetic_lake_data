from openpyxl import load_workbook
import numpy as np
from datetime import datetime
from Person import Person
from Account import Account, Transaction

# Началные свойства
n = 20000
quantity_person = 7000
wb = load_workbook("lake_data1.xlsx")
wb_name = load_workbook("data_names.xlsx")

ws1 = wb["A1"] #начальный отчет
ws2 = wb["A2"] #обогащенный набор

ws_personal_data = wb_name["names"] # набор = Ф.И.О, адрес, email

# Ф-я генерации персональных данных
def gen_person(quantity:int):
        arr_person = []
        for i in range(quantity):
                arr_person.append(Person())
        return arr_person

# Согласно генерации персональных данных создается аккаунты пользователей
def gen_account(persons):
        arr_account = []
        for person in persons:
                if np.random.randint(1,100) < 75:
                        arr_account.append([person,Account()])
                else:
                        for i in range(3):
                                arr_account.append([person,Account()])
        return arr_account

# Генерация транзакций пользователей
def gen_tansaction(row:int=20000,quantity_person = 7000):
        arr_persons = gen_person(quantity_person)
        arr_acc = gen_account(arr_persons)
        arr_tx =[]
        for i in range(2,row+2):
                num = np.random.randint(0,len(arr_acc))
                account = arr_acc[num][1]
                dt_in = datetime.strptime(account.acc_date,"%d/%m/%Y").strftime("%Y-%m-%d")
                dt_out = datetime.strptime(account.acc_close_date,"%d/%m/%Y").strftime("%Y-%m-%d")
                days = np.datetime64(dt_out)-np.datetime64(dt_in)
                txdate = np.datetime64(f"{dt_in}T00:00") + (np.random.randint(1,np.int64(days)) * np.random.randint(1,1440))
                tx = Transaction(txdate=str(txdate),currency=account.currency)
                arr_tx.append([arr_acc[num][0],arr_acc[num][1],tx,i])

        return arr_tx

# Генерация данных
tx = gen_tansaction(row=n, quantity_person=quantity_person)


# Запольнения данных в таблицу
def title_a1():

        for arg in tx:
                person = arg[0]
                account = arg[1]
                transaction = arg[2]
                row = arg[3]
                account.set_date(datetime.strptime(transaction.txdate, "%Y-%m-%dT%H:%M").strftime("%d/%m/%Y"))
                ws1["A" + str(row)] = row - 1                           # ID
                ws1["B" + str(row)] = person.lastname                   # Фамилия
                ws1["C" + str(row)] = person.firstname                  # Имя
                ws1["D" + str(row)] = person.thirdname                  # Отчество
                ws1["E" + str(row)] = person.birthday                   # День рождения
                ws1["F" + str(row)] = str(person.sex) + "`"             # Пол
                ws1["G" + str(row)] = str(person.marg_status) + "`"     # Статус
                ws1["H" + str(row)] = str(person.nationality) + "`"     # Национальность
                ws1["I" + str(row)] = str(person.snils) + "`"           # СНИЛС
                ws1["J" + str(row)] = str(person.born_place) + "`"      # Место рождения
                ws1["K" + str(row)] = str(account.account) + "`"        # Аккаунт
                ws1["L" + str(row)] = account.acc_date                  # Дата открытия
                ws1["M" + str(row)] = account.acc_close_date            # Дата закрытия
                ws1["N" + str(row)] = str(account.currency) + "`"       # Валюта
                ws1["O" + str(row)] = str(account.turn_per_day) + "`"   # Сред. кредит
                ws1["P" + str(row)] = str(account.acc_balance) + "`"    # Баланс
                ws1["Q" + str(row)] = account.acc_last_day              # Последний день оплаты
                ws1["R" + str(row)] = transaction.txdate                # Дата транзакции
                ws1["S" + str(row)] = str(transaction.txsum) + "`"      # Сумма транзакции
                ws1["T" + str(row)] = transaction.txplace               # Место транзакции
                # print(row)

# Обезличенный набор данных
def title_a2():

        for arg in tx:
                person = arg[0]
                account = arg[1]
                transaction = arg[2]
                row = arg[3]
                account.set_date(datetime.strptime(transaction.txdate, "%Y-%m-%dT%H:%M").strftime("%d/%m/%Y"))
                ws2["A" + str(row)] = row - 1                                   # ID
                ws2["B" + str(row)] = person.lastname[:2]                       # Фамилия
                ws2["C" + str(row)] = person.firstname[:1]                      # Имя
                ws2["D" + str(row)] = person.thirdname[:1]                      # Отчество
                ws2["E" + str(row)] = person.birthday[6:] + "`"                       # День рождения
                ws2["F" + str(row)] = str(person.sex ) + "`"                              # Пол
                ws2["G" + str(row)] = str(person.marg_status) + "`"                       # Статус
                ws2["H" + str(row)] = str(person.nationality) + "`"                        # Национальность
                ws2["I" + str(row)] = person.snils[:3] + "-xxx-xxx xx"          # СНИЛС
                ws2["J" + str(row)] = str(person.born_place) + "`"                     # Место рождения
                ws2["K" + str(row)] = str(account.account)[:5] + "xxxxxxx"      # Аккаунт
                ws2["L" + str(row)] = account.acc_date[3:] + "`"                  # Дата открытия
                ws2["M" + str(row)] = account.acc_close_date[3:] + "`"               # Дата закрытия
                ws2["N" + str(row)] = str(account.currency) + "`"                         # Валюта
                ws2["O" + str(row)] = str(round(account.turn_per_day + np.random.uniform(-2,2))) + "`"     # Сред. кредит
                ws2["P" + str(row)] = str(round(account.acc_balance + np.random.uniform(-2,2))) + "`"     # Баланс
                ws2["Q" + str(row)] = str(account.acc_last_day[3:])  + "`"                                # Последний день оплаты
                ws2["R" + str(row)] = transaction.txdate[:7]                                    # Дата транзакции
                ws2["S" + str(row)] = str(round(transaction.txsum + np.random.uniform(-2,2))) +"`"        # Сумма транзакции
                ws2["T" + str(row)] = transaction.txplace                                       # Место транзакции
                print(row)


title_a1()
title_a2()
wb.save(filename="lake_data_v.0.10.xlsx")
